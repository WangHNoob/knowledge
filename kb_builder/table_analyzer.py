# -*- coding: utf-8 -*-
"""
Deterministic xlsx table analysis (no LLM).

Responsibilities:
  1. Scan every *.xlsx under knowledge/gamedata/, read header row of the first
     sheet, produce per-table schema (field list + path + group).
  2. Group tables into "families" by folder or PascalCase prefix (borrowed
     from old/wiki_compiler.py::_extract_table_groups).
  3. Convention-based FK detection: a field ending in `Id`/`ID` whose stem
     (case-insensitive) matches another table's name becomes a FK edge.
  4. Generate wiki/tables/<family>.md (one page per family).
  5. Serialize machine-readable registries:
       knowledge/wiki/_tables/schemas.json        (all table schemas)
       knowledge/wiki/_tables/groups.json         (family -> [tables])
       knowledge/wiki/_tables/table_fk_registry.json (all FK edges)

Caching: (rel_path, mtime, size) → fields; on subsequent runs unchanged
xlsx files are not re-opened. First full scan of 7500+ files takes minutes.
"""
from __future__ import annotations

import argparse
import collections
import hashlib
import json
import os
import re
import time
from typing import Optional

from .config import PATHS

PROJECT_ROOT = str(PATHS.project_root)
GAMEDATA_DIR = str(PATHS.gamedata_dir)
WIKI_DIR = str(PATHS.wiki_dir)
TABLES_DIR = str(PATHS.wiki_tables_dir)
REGISTRY_DIR = str(PATHS.wiki_tables_registry_dir)
SCHEMAS_PATH = str(PATHS.schemas_json_path)
GROUPS_PATH = str(PATHS.groups_json_path)
FK_PATH = str(PATHS.fk_registry_json_path)

HEADER_ROW_CANDIDATES = (1, 2, 3)  # 策划表常把第 1 行用作中文注释，第 2/3 行才是字段名
MIN_FIELDS = 2  # 少于这个数的 header 当无效


# =================================================================
# 扫描：读 xlsx header
# =================================================================

def _read_header_openpyxl(path: str) -> list[str]:
    """读取 xlsx 第一个 visible sheet 的前几行，挑出最像字段名的一行。"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl not installed")

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return []

    try:
        # 选第一个可见 sheet
        sheet = None
        for ws in wb.worksheets:
            if getattr(ws, "sheet_state", "visible") == "visible":
                sheet = ws
                break
        if sheet is None and wb.worksheets:
            sheet = wb.worksheets[0]
        if sheet is None:
            return []

        # 取前 3 行
        rows: list[list[str]] = []
        for i, row in enumerate(sheet.iter_rows(values_only=True, max_row=3), start=1):
            rows.append([
                ("" if v is None else str(v)).strip()
                for v in row
            ])
            if i >= 3:
                break

        # 选列数最多且字段看起来像 identifier 的行
        best: list[str] = []
        best_score = -1
        for r in rows:
            non_empty = [c for c in r if c]
            if len(non_empty) < MIN_FIELDS:
                continue
            # 看像 identifier (英文+数字+下划线) 的比例
            id_like = sum(1 for c in non_empty if re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", c))
            score = id_like * 2 + len(non_empty)
            if score > best_score:
                best_score = score
                best = non_empty
        return best
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _load_schema_cache() -> dict:
    if not os.path.exists(SCHEMAS_PATH):
        return {}
    try:
        with open(SCHEMAS_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def scan_schemas(gamedata_dir: str = GAMEDATA_DIR,
                 force: bool = False,
                 progress_every: int = 200) -> dict:
    """扫描 gamedata 下所有 xlsx，返回 {table_name: {fields, rel_path, size, mtime}}。

    带 (rel_path, mtime, size) 指纹缓存，未变更文件直接复用。
    """
    if not os.path.isdir(gamedata_dir):
        print(f"[warn] {gamedata_dir} not found")
        return {}

    old_cache = {} if force else _load_schema_cache()

    # 收集所有 xlsx
    entries: list[tuple[str, str]] = []  # (abs_path, rel_path)
    for root, _dirs, files in os.walk(gamedata_dir):
        for fn in files:
            if fn.lower().endswith(".xlsx") and not fn.startswith("~"):
                ap = os.path.join(root, fn)
                rp = os.path.relpath(ap, gamedata_dir).replace(os.sep, "/")
                entries.append((ap, rp))
    entries.sort(key=lambda x: x[1])
    print(f"发现 {len(entries)} 个 xlsx 文件")

    schemas: dict[str, dict] = {}
    reused = 0
    parsed = 0
    skipped = 0
    t0 = time.time()

    for i, (abs_path, rel_path) in enumerate(entries, start=1):
        try:
            st = os.stat(abs_path)
        except OSError:
            skipped += 1
            continue
        mtime = int(st.st_mtime)
        size = st.st_size
        table_name = _table_name_from_rel(rel_path)

        cached = old_cache.get(table_name)
        if (cached and cached.get("rel_path") == rel_path
                and cached.get("mtime") == mtime
                and cached.get("size") == size):
            schemas[table_name] = cached
            reused += 1
        else:
            fields = _read_header_openpyxl(abs_path)
            schemas[table_name] = {
                "rel_path": rel_path,
                "mtime": mtime,
                "size": size,
                "fields": fields,
            }
            parsed += 1

        if i % progress_every == 0 or i == len(entries):
            dt = time.time() - t0
            rate = i / dt if dt else 0
            print(f"  进度 {i}/{len(entries)}  reused={reused} parsed={parsed}  ({rate:.0f}/s)")

    # 记录分组（基于全量表名）
    groups = _group_tables(schemas)
    for tname, info in schemas.items():
        info["group"] = _find_group(tname, groups)

    print(f"扫描完成：reused {reused}, parsed {parsed}, skipped {skipped}")
    return schemas


def _table_name_from_rel(rel_path: str) -> str:
    """knowledge/gamedata/<rel> 中的相对路径 → 表名。

    子目录的表名带路径前缀：'config/Meta' 而不是 'Meta'。
    """
    stem = rel_path[:-5] if rel_path.lower().endswith(".xlsx") else rel_path
    return stem


# =================================================================
# 分组：文件夹 + PascalCase 前缀
# =================================================================

def _group_tables(schemas: dict) -> dict[str, list[str]]:
    """参考 old/wiki_compiler.py::_extract_table_groups。"""
    folder_groups: dict[str, list[str]] = collections.defaultdict(list)
    root_tables: list[str] = []

    for tname in schemas:
        parts = tname.split("/")
        if len(parts) > 1:
            folder_groups[parts[0]].append(tname)
        else:
            root_tables.append(tname)

    prefix_groups: dict[str, list[str]] = collections.defaultdict(list)
    for tname in root_tables:
        # PascalCase 首词；若以 _ 开头 (如 _Buff) 取 _Xxx
        m = re.match(r"(_?[A-Z][a-z0-9]*)", tname)
        prefix = m.group(1) if m else tname
        if len(prefix.lstrip("_")) >= 2:
            prefix_groups[prefix].append(tname)

    groups: dict[str, list[str]] = dict(folder_groups)
    for prefix, tables in prefix_groups.items():
        if len(tables) >= 2 and prefix not in groups:
            groups[prefix] = tables

    # 零散 singleton 收到 "_misc" 组
    grouped_set: set[str] = set()
    for lst in groups.values():
        grouped_set.update(lst)
    misc = [t for t in root_tables if t not in grouped_set]
    if misc:
        groups["_misc"] = sorted(misc)

    # 每组内部排序
    for k in groups:
        groups[k] = sorted(set(groups[k]))
    return groups


def _find_group(tname: str, groups: dict[str, list[str]]) -> str:
    for g, members in groups.items():
        if tname in members:
            return g
    return "_misc"


# =================================================================
# FK 检测：字段命名约定
# =================================================================

FK_FIELD_RE = re.compile(r"^(.+?)[_]?[Ii][Dd]s?$")  # BuffId / Buff_Id / BuffIDs


def detect_fk_edges(schemas: dict) -> list[dict]:
    """检测表间外键。

    规则：字段名形如 `<Stem>Id(s)` 且 <Stem> 不区分大小写匹配到另一张表名
    （或该表名去掉前置下划线后的形式）时，产生一条边。
    """
    # 表名标准化：去掉路径前缀 + 去掉前置下划线；小写键 → 原表名
    name_index: dict[str, str] = {}
    for tname in schemas:
        simple = tname.split("/")[-1].lstrip("_")
        name_index[simple.lower()] = tname

    edges: list[dict] = []
    seen: set[tuple[str, str, str]] = set()
    for tname, info in schemas.items():
        my_simple = tname.split("/")[-1].lstrip("_").lower()
        for field in info.get("fields") or []:
            m = FK_FIELD_RE.match(field)
            if not m:
                continue
            stem = m.group(1)
            if not stem or stem.lower() == my_simple:
                continue
            target = name_index.get(stem.lower())
            if not target or target == tname:
                continue
            key = (tname, target, field)
            if key in seen:
                continue
            seen.add(key)
            edges.append({
                "source": tname, "target": target,
                "field": field, "source_of_edge": "field_convention",
            })
    edges.sort(key=lambda e: (e["source"], e["target"], e["field"]))
    return edges


# =================================================================
# 输出：wiki pages + registry JSON
# =================================================================

def write_registry_files(schemas: dict, groups: dict[str, list[str]],
                          fk_edges: list[dict]) -> None:
    os.makedirs(REGISTRY_DIR, exist_ok=True)

    # schemas: sort keys for determinism
    serializable_schemas = {k: schemas[k] for k in sorted(schemas)}
    with open(SCHEMAS_PATH, "w", encoding="utf-8") as f:
        json.dump(serializable_schemas, f, ensure_ascii=False, indent=2, sort_keys=True)

    with open(GROUPS_PATH, "w", encoding="utf-8") as f:
        json.dump({k: groups[k] for k in sorted(groups)},
                  f, ensure_ascii=False, indent=2)

    with open(FK_PATH, "w", encoding="utf-8") as f:
        json.dump(fk_edges, f, ensure_ascii=False, indent=2)

    print(f"  写 {SCHEMAS_PATH} ({len(schemas)} tables)")
    print(f"  写 {GROUPS_PATH} ({len(groups)} groups)")
    print(f"  写 {FK_PATH} ({len(fk_edges)} FK edges)")


def write_table_pages(schemas: dict, groups: dict[str, list[str]],
                      fk_edges: list[dict]) -> None:
    os.makedirs(TABLES_DIR, exist_ok=True)

    # 预计算每张表的出/入 FK
    out_fk: dict[str, list[dict]] = collections.defaultdict(list)
    in_fk: dict[str, list[dict]] = collections.defaultdict(list)
    for e in fk_edges:
        out_fk[e["source"]].append(e)
        in_fk[e["target"]].append(e)

    # 为避免与一族一页冲突：所有文件名必须安全
    for group_name, tables in groups.items():
        slug = _slugify_group(group_name)
        path = os.path.join(TABLES_DIR, f"{slug}.md")
        lines: list[str] = []
        lines.append("---")
        lines.append("type: table_schema")
        lines.append(f'title: "表族 {group_name}"')
        lines.append(f'group: "{group_name}"')
        lines.append(f"table_count: {len(tables)}")
        lines.append("---")
        lines.append("")
        lines.append(f"# 表族 `{group_name}`")
        lines.append("")
        lines.append(f"共 {len(tables)} 张表。数据源：`knowledge/gamedata/`。")
        lines.append("")
        lines.append("## 成员表清单")
        lines.append("| 表名 | 字段数 | 相对路径 |")
        lines.append("|------|--------|----------|")
        for t in tables:
            info = schemas.get(t, {})
            lines.append(
                f"| `{t}` | {len(info.get('fields') or [])} | {info.get('rel_path', '?')} |"
            )
        lines.append("")

        lines.append("## 字段明细")
        for t in tables:
            info = schemas.get(t, {})
            fields = info.get("fields") or []
            lines.append(f"### `{t}`")
            if not fields:
                lines.append("_未读取到字段（文件可能为空或 header 解析失败）_")
                lines.append("")
                continue
            # 字段按每行 8 个排版
            chunk = 8
            for i in range(0, len(fields), chunk):
                row = fields[i:i + chunk]
                lines.append("- " + " · ".join(f"`{f}`" for f in row))
            # 出/入 FK
            of = out_fk.get(t, [])
            inf = in_fk.get(t, [])
            if of:
                lines.append("")
                lines.append("**出向外键** ({}):".format(len(of)))
                for e in of[:20]:
                    lines.append(f"- `{e['field']}` → `{e['target']}`")
                if len(of) > 20:
                    lines.append(f"- … 其余 {len(of) - 20} 条见 `_tables/table_fk_registry.json`")
            if inf:
                lines.append("")
                lines.append("**入向外键** ({}):".format(len(inf)))
                for e in inf[:10]:
                    lines.append(f"- `{e['source']}.{e['field']}` → 本表")
                if len(inf) > 10:
                    lines.append(f"- … 其余 {len(inf) - 10} 条见 `_tables/table_fk_registry.json`")
            lines.append("")

        with open(path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))

    print(f"  写 {len(groups)} 份族页 -> {TABLES_DIR}/")


def _slugify_group(name: str) -> str:
    s = re.sub(r"[^\w.-]+", "_", name).strip("_")
    return s or "_group"


# =================================================================
# 入口
# =================================================================

def run(force: bool = False) -> dict:
    schemas = scan_schemas(force=force)
    if not schemas:
        print("[warn] no schemas scanned, skip")
        return {}
    groups = _group_tables(schemas)
    fk_edges = detect_fk_edges(schemas)
    print(f"分组: {len(groups)} 族；FK 边: {len(fk_edges)}")
    write_registry_files(schemas, groups, fk_edges)
    write_table_pages(schemas, groups, fk_edges)
    return {"schemas": schemas, "groups": groups, "fk_edges": fk_edges}


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Deterministic xlsx table analysis")
    parser.add_argument("-f", "--force", action="store_true",
                        help="忽略缓存，重新读取所有 xlsx 头部")
    args = parser.parse_args()
    run(force=args.force)
