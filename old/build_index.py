# -*- coding: utf-8 -*-
"""
大表 SQLite 索引构建器 -- 对超过阈值的大 xlsx 文件建立结构化索引。

每张表独立一个 SQLite table，列名即字段名，保留原始表结构。
同时为每张表创建 FTS5 全文搜索虚拟表（trigram 分词器，支持中文子串匹配）。

用法：
    python build_index.py                    # 增量构建
    python build_index.py -f                 # 强制全量重建
    python build_index.py --threshold 500    # 自定义大小阈值 (KB)
    python build_index.py --query "龙泉剑"   # 测试全文检索

LLM 检索示例（SQL）：
    -- 查看表结构
    PRAGMA table_info(NpcTalk)
    -- 精确查询
    SELECT * FROM NpcTalk WHERE NpcName LIKE '%主角%'
    -- 全文搜索
    SELECT * FROM fts_NpcTalk WHERE fts_NpcTalk MATCH '龙泉剑'
"""

import os
import sys

# 加载 .env 环境变量
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass
import re
import json
import sqlite3
import hashlib
import argparse
import time
from pathlib import Path
from datetime import datetime, timezone

# ══════════════════════════════════════════════════════════════
# 配置
# ══════════════════════════════════════════════════════════════

# 默认大表阈值（字节），文件大小 >= 此值才建索引
DEFAULT_THRESHOLD_KB = 200
DEFAULT_THRESHOLD = DEFAULT_THRESHOLD_KB * 1024

# 输出文件
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_INDEX_DB = os.path.join(SCRIPT_DIR, 'knowledge', 'index.db')

# gamedata 目录
DEFAULT_GAMEDATA_DIR = os.path.join(SCRIPT_DIR, 'knowledge', 'gamedata')


# ══════════════════════════════════════════════════════════════
# Schema
# ══════════════════════════════════════════════════════════════

_META_SCHEMA = """
-- 构建元数据
CREATE TABLE IF NOT EXISTS _meta (
    key   TEXT PRIMARY KEY,
    value TEXT NOT NULL
);

-- 每张被索引的表的注册信息
CREATE TABLE IF NOT EXISTS _tables (
    table_name  TEXT PRIMARY KEY,
    clean_name  TEXT NOT NULL,       -- SQLite 安全表名
    file_path   TEXT NOT NULL,
    file_size   INTEGER NOT NULL,
    row_count   INTEGER NOT NULL,
    col_names   TEXT NOT NULL,       -- JSON array: ["id", "name", ...]
    col_types   TEXT NOT NULL,       -- JSON array: ["int", "string", ...]
    sheet_names TEXT NOT NULL,       -- JSON array: ["Sheet1", "Sheet2", ...]
    file_hash   TEXT NOT NULL,       -- sha256 前16位，用于增量检测
    indexed_at  TEXT NOT NULL        -- ISO timestamp
);
"""


def _clean_table_name(name):
    """将表名转为 SQLite 安全标识符（只保留字母数字下划线）。"""
    return re.sub(r'[^a-zA-Z0-9_]', '_', name)


def _create_table_sql(clean_name, col_names, col_types):
    """生成创建数据表的 SQL。

    每张表有 _rowid INTEGER PRIMARY KEY 自动递增，其余列为 TEXT 类型
    （保持数据完整性，避免类型转换丢失信息）。
    """
    cols = ['_rowid INTEGER PRIMARY KEY AUTOINCREMENT']
    for i, col in enumerate(col_names):
        # SQLite 列名也需要清理
        safe_col = re.sub(r'[^a-zA-Z0-9_\u4e00-\u9fff]', '_', col)
        # 去掉前导数字
        if safe_col and safe_col[0].isdigit():
            safe_col = '_' + safe_col
        # 空列名用占位符
        if not safe_col or safe_col == '_':
            safe_col = f'_col{i}'
        # 去重
        cols.append(f'[{safe_col}] TEXT')
    return f"CREATE TABLE IF NOT EXISTS [{clean_name}] ({', '.join(cols)})"


def _create_fts_sql(clean_name, col_names):
    """生成创建 FTS5 全文搜索虚拟表的 SQL。"""
    # FTS5 不支持中文列名，用 c0, c1, ... 作为 FTS 列名
    fts_name = f'fts_{clean_name}'
    fts_cols = ['c{}'.format(i) for i in range(len(col_names))]
    cols_def = ', '.join(fts_cols)
    return (
        f"CREATE VIRTUAL TABLE IF NOT EXISTS [{fts_name}] USING fts5("
        f"{cols_def}, tokenize='trigram')"
    )


# ══════════════════════════════════════════════════════════════
# 工具函数
# ══════════════════════════════════════════════════════════════

def _file_hash(filepath):
    """计算文件 sha256 前16位（快速变更检测）。"""
    h = hashlib.sha256()
    with open(filepath, 'rb') as f:
        # 只读前 64KB + 最后 64KB + 文件大小，平衡速度和准确性
        size = os.path.getsize(filepath)
        h.update(size.to_bytes(8, 'little'))
        if size > 128 * 1024:
            f.seek(0)
            h.update(f.read(64 * 1024))
            f.seek(-64 * 1024, 2)
            h.update(f.read(64 * 1024))
        else:
            h.update(f.read())
    return h.hexdigest()[:16]


def _clean_cell(value):
    """清理单元格值：NaN → None，去除首尾空白。"""
    if value is None:
        return None
    s = str(value).strip()
    if s in ('', 'NaN', 'nan', 'None', '#REF!', '#N/A', '#VALUE!'):
        return None
    return s


def _scan_xlsx_files(gamedata_dir, threshold):
    """扫描目录，返回超过阈值的大 xlsx 文件列表。

    Returns:
        list[tuple[str, str, int]]: [(table_name, abs_path, file_size), ...]
    """
    results = []
    if not os.path.isdir(gamedata_dir):
        return results

    for root, _dirs, files in os.walk(gamedata_dir):
        for fname in files:
            if fname.lower().endswith('.xlsx') and not fname.startswith('~'):
                fpath = os.path.join(root, fname)
                fsize = os.path.getsize(fpath)
                if fsize >= threshold:
                    rel = os.path.relpath(root, gamedata_dir)
                    stem = fname[:-5]
                    table_name = stem if rel == '.' else rel.replace(os.sep, '/') + '/' + stem
                    results.append((table_name, fpath, fsize))

    return results


def _read_xlsx_table(filepath):
    """用 openpyxl 读取 xlsx 的所有 sheet 数据。

    Returns:
        tuple: (headers, rows, sheet_info)
            headers: list[str] - 合并后的列名
            rows: list[list] - 每行一个值列表（顺序与 headers 对应）
            sheet_info: list[str] - 所有 sheet 名
    """
    import openpyxl

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    all_sheets = wb.sheetnames
    all_rows = []
    best_headers = []
    best_header_count = 0
    header_idx = 0

    for sheet_name in all_sheets:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # 检测表头：前10行中非空值最多的行
        for i, row in enumerate(rows[:10]):
            non_none = sum(1 for c in row if c is not None and str(c).strip() not in ('', 'NaN', 'nan'))
            if non_none > best_header_count:
                best_header_count = non_none
                best_headers = list(row)
                header_idx = i
                if non_none >= 3:
                    break

        # 从表头下一行开始取数据
        col_names = []
        for i, h in enumerate(best_headers):
            cleaned = _clean_cell(h)
            col_names.append(cleaned if cleaned else f'_col{i}')

        for row in rows[header_idx + 1:]:
            if all(c is None for c in row):
                continue
            record = []
            for i in range(len(col_names)):
                val = row[i] if i < len(row) else None
                record.append(_clean_cell(val))
            all_rows.append(record)

    wb.close()

    if not best_headers:
        return [], [], all_sheets

    # 去重列名（同名列加后缀）
    seen = {}
    final_headers = []
    for h in best_headers:
        cleaned = _clean_cell(h) if h else None
        name = cleaned if cleaned else '_col'
        if name in seen:
            seen[name] += 1
            name = f'{name}_{seen[name]}'
        else:
            seen[name] = 0
        final_headers.append(name)

    return final_headers, all_rows, all_sheets


# ══════════════════════════════════════════════════════════════
# 索引构建
# ══════════════════════════════════════════════════════════════

def build_index(gamedata_dir, db_path, threshold=DEFAULT_THRESHOLD, force=False):
    """构建/增量更新 SQLite 索引。

    每张表独立一个 SQLite table，列名即字段名。
    同时为每张表创建 fts_{name} FTS5 全文搜索虚拟表。

    Args:
        gamedata_dir: knowledge/gamedata/ 目录路径
        db_path: SQLite 数据库路径
        threshold: 文件大小阈值（字节），超过此值才建索引
        force: 是否强制全量重建
    """
    # 扫描大表
    big_files = _scan_xlsx_files(gamedata_dir, threshold)
    big_files.sort(key=lambda x: x[0])  # 按表名排序
    print(f"[i] 大表阈值: {threshold // 1024} KB")
    print(f"[i] 扫描到 {len(big_files)} 张大表需要索引\n")

    if not big_files:
        print("[i] 没有大表需要索引，退出。")
        return

    # 初始化数据库
    os.makedirs(os.path.dirname(db_path), exist_ok=True)
    conn = sqlite3.connect(db_path)
    conn.executescript(_META_SCHEMA)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")

    # 读取已索引的表的 hash
    existing = {}
    if not force:
        for row in conn.execute("SELECT table_name, file_hash FROM _tables"):
            existing[row[0]] = row[1]

    # 统计
    indexed = 0
    skipped = 0
    failed = 0
    total_rows = 0

    for table_name, filepath, file_size in big_files:
        # 增量检测：hash 未变则跳过
        fhash = _file_hash(filepath)
        if not force and existing.get(table_name) == fhash:
            skipped += 1
            continue

        clean = _clean_table_name(table_name)

        try:
            headers, rows, sheet_names = _read_xlsx_table(filepath)
            if not rows:
                print(f"  [SKIP] {table_name}: 空表")
                skipped += 1
                continue

            # 推断列类型
            col_types = []
            for i, h in enumerate(headers):
                vals = [r[i] for r in rows[:100] if i < len(r) and r[i] is not None]
                if not vals:
                    col_types.append('unknown')
                elif all(re.match(r'^-?\d+$', str(v)) for v in vals):
                    col_types.append('int')
                elif all(re.match(r'^-?\d+(\.\d+)?$', str(v)) for v in vals):
                    col_types.append('float')
                else:
                    col_types.append('string')

            now = datetime.now(timezone.utc).isoformat()

            # 删除旧表和旧 FTS
            try:
                conn.execute(f"DROP TABLE IF EXISTS [{clean}]")
            except Exception:
                pass
            try:
                conn.execute(f"DROP TABLE IF EXISTS [fts_{clean}]")
            except Exception:
                pass
            conn.execute("DELETE FROM _tables WHERE table_name = ?", (table_name,))

            # 创建数据表
            conn.execute(_create_table_sql(clean, headers, col_types))

            # 创建 FTS5 全文搜索表
            conn.execute(_create_fts_sql(clean, headers))

            # 生成安全列名列表
            safe_cols = []
            for i, col in enumerate(headers):
                safe_col = re.sub(r'[^a-zA-Z0-9_\u4e00-\u9fff]', '_', col)
                if safe_col and safe_col[0].isdigit():
                    safe_col = '_' + safe_col
                if not safe_col or safe_col == '_':
                    safe_col = f'_col{i}'
                safe_cols.append(f'[{safe_col}]')

            col_names_sql = ', '.join(safe_cols)
            placeholders = ', '.join(['?'] * len(headers))

            # FTS 列名: c0, c1, ...
            fts_cols = ', '.join([f'c{i}' for i in range(len(headers))])
            fts_placeholders = ', '.join(['?'] * len(headers))

            # 批量插入行（每 1000 行一个事务）
            batch_data = []
            batch_fts = []
            n_cols = len(headers)

            for row_vals in rows:
                # 补齐/截断列数
                padded = list(row_vals) + [None] * max(0, n_cols - len(row_vals))
                data_row = padded[:n_cols]
                # 数据表：转为字符串（TEXT 列）
                str_row = [str(v) if v is not None else None for v in data_row]
                batch_data.append(str_row)
                # FTS 表：所有值拼成字符串用于全文搜索
                batch_fts.append(str_row)

                if len(batch_data) >= 1000:
                    conn.executemany(
                        f"INSERT INTO [{clean}] ({col_names_sql}) VALUES ({placeholders})",
                        batch_data
                    )
                    conn.executemany(
                        f"INSERT INTO [fts_{clean}] ({fts_cols}) VALUES ({fts_placeholders})",
                        batch_fts
                    )
                    conn.commit()
                    total_rows += len(batch_data)
                    batch_data = []
                    batch_fts = []

            if batch_data:
                conn.executemany(
                    f"INSERT INTO [{clean}] ({col_names_sql}) VALUES ({placeholders})",
                    batch_data
                )
                conn.executemany(
                    f"INSERT INTO [fts_{clean}] ({fts_cols}) VALUES ({fts_placeholders})",
                    batch_fts
                )
                conn.commit()
                total_rows += len(batch_data)

            # 写入表元数据
            conn.execute(
                """INSERT INTO _tables (table_name, clean_name, file_path, file_size, row_count,
                   col_names, col_types, sheet_names, file_hash, indexed_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (table_name,
                 clean,
                 os.path.relpath(filepath, SCRIPT_DIR),
                 file_size,
                 len(rows),
                 json.dumps(headers, ensure_ascii=False),
                 json.dumps(col_types, ensure_ascii=False),
                 json.dumps(sheet_names, ensure_ascii=False),
                 fhash,
                 now)
            )
            conn.commit()

            indexed += 1
            size_kb = file_size / 1024
            print(f"  [OK] {table_name}: {len(rows):,} rows, {len(headers)} cols ({size_kb:,.0f} KB)")

        except Exception as e:
            failed += 1
            print(f"  [ERR] {table_name}: {e}")

    # 写入构建元数据
    now = datetime.now(timezone.utc).isoformat()
    conn.execute(
        "INSERT OR REPLACE INTO _meta (key, value) VALUES (?, ?)",
        ('build_info', json.dumps({
            'built_at': now,
            'threshold_kb': threshold // 1024,
            'total_tables': len(big_files),
            'indexed': indexed,
            'skipped': skipped,
            'failed': failed,
            'total_rows': total_rows,
        }, ensure_ascii=False))
    )
    conn.commit()
    conn.close()

    # 输出统计
    db_size = os.path.getsize(db_path) / 1024 / 1024
    print(f"\n{'='*50}")
    print(f"[OK] 索引构建完成:")
    print(f"     新建: {indexed}, 跳过: {skipped}, 失败: {failed}")
    print(f"     总行数: {total_rows:,}")
    print(f"     数据库: {db_path} ({db_size:.1f} MB)")
    print(f"{'='*50}")


def query_index(db_path, query_text, table_name=None, limit=10):
    """测试全文检索。

    支持两种模式：
    1. 指定 table_name: 在 fts_{table} 中搜索
    2. 不指定: 遍历所有 fts_{table} 搜索

    Args:
        db_path: SQLite 数据库路径
        query_text: 搜索文本
        table_name: 可选，限定表名
        limit: 返回条数
    """
    if not os.path.exists(db_path):
        print(f"[ERR] 数据库不存在: {db_path}")
        print("  请先运行: python build_index.py")
        return

    conn = sqlite3.connect(db_path)

    # 显示已索引的表
    tables = conn.execute(
        "SELECT table_name, clean_name, row_count, col_names FROM _tables ORDER BY row_count DESC"
    ).fetchall()
    print(f"已索引 {len(tables)} 张表:\n")
    for t, cn, cnt, cols in tables[:10]:
        col_list = json.loads(cols)
        print(f"  {t} (→ {cn}): {cnt:,} rows, cols={col_list[:6]}{'...' if len(col_list)>6 else ''}")
    if len(tables) > 10:
        print(f"  ... 还有 {len(tables)-10} 张表")

    # trigram 搜索：空格分隔的词用 AND 连接
    terms = query_text.strip().split()
    fts_query = ' AND '.join(terms)

    print(f"\n搜索: \"{query_text}\" (FTS5: \"{fts_query}\")\n")

    if table_name:
        # 指定表名搜索
        # 找到 clean_name
        clean = None
        for t, cn, cnt, cols in tables:
            if t == table_name or cn == table_name:
                clean = cn
                break
        if not clean:
            print(f"  [ERR] 表 {table_name} 未找到")
            conn.close()
            return

        fts_table = f'fts_{clean}'
        # 获取列名
        col_info = conn.execute(f"PRAGMA table_info([{clean}])").fetchall()
        col_names = [c[1] for c in col_info if c[1] != '_rowid']

        try:
            sql = f"SELECT _rowid, * FROM [{fts_table}] WHERE [{fts_table}] MATCH ? LIMIT ?"
            rows = conn.execute(sql, (fts_query, limit)).fetchall()
        except Exception as e:
            print(f"  [ERR] 搜索失败: {e}")
            conn.close()
            return

        if not rows:
            print("  (无匹配结果)")
        else:
            for row in rows:
                rowid = row[0]
                values = row[1:]
                preview = {}
                for i, v in enumerate(values):
                    if i < len(col_names) and v:
                        preview[col_names[i]] = str(v)[:100]
                print(f"  [{table_name} row#{rowid}]")
                print(f"  {json.dumps(preview, ensure_ascii=False, indent=2)[:500]}")
                print()
    else:
        # 遍历所有表搜索
        found = 0
        for t, cn, cnt, cols_json in tables:
            if found >= limit:
                break
            fts_table = f'fts_{cn}'
            col_info = conn.execute(f"PRAGMA table_info([{cn}])").fetchall()
            col_names = [c[1] for c in col_info if c[1] != '_rowid']

            try:
                sql = f"SELECT _rowid, * FROM [{fts_table}] WHERE [{fts_table}] MATCH ? LIMIT ?"
                rows = conn.execute(sql, (fts_query, limit)).fetchall()
            except Exception:
                continue

            for row in rows:
                if found >= limit:
                    break
                rowid = row[0]
                values = row[1:]
                preview = {}
                for i, v in enumerate(values):
                    if i < len(col_names) and v:
                        preview[col_names[i]] = str(v)[:100]
                print(f"  [{t} row#{rowid}]")
                print(f"  {json.dumps(preview, ensure_ascii=False, indent=2)[:500]}")
                print()
                found += 1

        if not found:
            print("  (无匹配结果)")

    conn.close()


# ══════════════════════════════════════════════════════════════
# CLI
# ══════════════════════════════════════════════════════════════

if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='大表 SQLite 索引构建器',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python build_index.py                       # 增量构建
  python build_index.py -f                    # 强制全量重建
  python build_index.py --threshold 500       # 只索引 >= 500KB 的表
  python build_index.py --query "龙泉剑"      # 测试全文检索
  python build_index.py --query "3000001" --table NpcTalk

SQL 查询示例:
  -- 查看表结构
  PRAGMA table_info(NpcTalk)
  -- 精确查询
  SELECT * FROM NpcTalk WHERE NpcName LIKE '%主角%'
  -- 全文搜索
  SELECT * FROM fts_NpcTalk WHERE fts_NpcTalk MATCH '龙泉剑'
        """)
    parser.add_argument('-f', '--force', action='store_true',
                        help='强制全量重建索引')
    parser.add_argument('--threshold', type=int, default=DEFAULT_THRESHOLD_KB,
                        help=f'大表阈值 KB (默认: {DEFAULT_THRESHOLD_KB})')
    parser.add_argument('--query', type=str, default=None,
                        help='测试全文检索查询')
    parser.add_argument('--table', type=str, default=None,
                        help='检索时限定表名')
    parser.add_argument('--db', type=str, default=DEFAULT_INDEX_DB,
                        help=f'数据库路径 (默认: knowledge/index.db)')
    parser.add_argument('--gamedata', type=str, default=DEFAULT_GAMEDATA_DIR,
                        help=f'gamedata 目录路径 (默认: knowledge/gamedata)')

    args = parser.parse_args()

    if args.query:
        query_index(args.db, args.query, args.table)
    else:
        t0 = time.time()
        build_index(
            args.gamedata,
            args.db,
            threshold=args.threshold * 1024,
            force=args.force
        )
        elapsed = time.time() - t0
        print(f"\n耗时: {elapsed:.1f}s")
